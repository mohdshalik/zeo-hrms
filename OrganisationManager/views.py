from django.shortcuts import render
from rest_framework.response import Response
from django.http import FileResponse, Http404
from rest_framework import status,generics,viewsets,permissions
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import api_view
from django.http import FileResponse
from openpyxl import load_workbook
from docx import Document
from django_tenants.utils import schema_context
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from .models import (brnch_mstr,dept_master,DocumentNumbering,
                     desgntn_master,ctgry_master,FiscalPeriod,FiscalYear,CompanyPolicy,Announcement,
                     AnnouncementComment,AnnouncementView,Asset, AssetAllocation,AssetRequest,AssetCustomField,AssetType,
                     AssetReport,AssetCustomFieldValue,AssetTransactionReport)

from . serializer import (BranchSerializer,PermissionSerializer,GroupSerializer,permserializer,DocumentNumberingSerializer,
                          CtgrySerializer,DeptSerializer,DesgSerializer,FiscalYearSerializer,PeriodSerializer,DeptUploadSerializer,CtgryUploadSerializer,
                          DesgUploadSerializer,CompanyPolicySerializer,AnnouncementSerializer,AnnouncementCommentSerializer,AssetSerializer,AssetAllocationSerializer,AssetRequestSerializer,AssetCustomFieldSerializer,
                          AssetTypeSerializer,AssetCustomFieldValueSerializer,AssetReportSerializer,AssetTransactionReportSerializer)
from rest_framework.permissions import IsAuthenticated,AllowAny,IsAuthenticatedOrReadOnly,IsAdminUser
from .resource import (DepartmentResource,DesignationResource,DesgtnReportResource,DeptReportResource,CategoryResource)
from EmpManagement.models import emp_master
from UserManagement.permissions import IsAdminUser,IsSuperUser
from datetime import timedelta
from rest_framework.views import APIView
from openpyxl import Workbook
from django.http import JsonResponse
from tablib import Dataset
from openpyxl.styles import PatternFill,Alignment,Font
from django.http import HttpResponse
from io import BytesIO
from .permissions import (BranchPermission,DepartmentPermission,DesignationPermission,CategoryPermission,FiscalYearPermission,DocumentNumberingPermission,
                          CompanyPolicyPermission,AssetMasterPermission,Asset_CustomFieldValuePermission,AssetTransactionPermission)
# Create your views here.
from django.contrib.auth.models import Permission,Group
from django.core.exceptions import PermissionDenied
from tenant_users.tenants.models import UserTenantPermissions
from rest_framework.decorators import action
import subprocess
import os
from django.utils import timezone
from django.db.models import Q,Field
from django.contrib.contenttypes.models import ContentType
from django.core.mail import EmailMessage
from EmpManagement.models import EmailConfiguration  # Adjust path
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.conf import settings
import json
from datetime import date,datetime
from django.core.cache import cache
from django.utils import timezone


def get_model_permissions(model):
    content_type = ContentType.objects.get_for_model(model)
    permissions = Permission.objects.filter(
        Q(content_type=content_type)
    )
    return permissions
class BranchViewSet(viewsets.ModelViewSet):
    queryset = brnch_mstr.objects.all()
    serializer_class = BranchSerializer
    permission_classes = [BranchPermission]
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['tenant_id'] = self.request.query_params.get('tenant_id')
        return context  
    
    @action(detail=True, methods=['get'])
    def departments(self, request, pk=None):
        branch = self.get_object()
        depts = branch.dept_master_set.all()  # Use the correct related_name
        serializer = DeptSerializer(depts, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['POST', 'GET'])
    def companypolicies(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            family_members = employee.policies.all()
            serializer = CompanyPolicySerializer(family_members, many=True)
            return Response(serializer.data)
#DEPARTMENT 
class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = dept_master.objects.all()
    serializer_class = DeptSerializer
    # authentication_classes = [SessionAuthentication,]
    permission_classes = [DepartmentPermission]
    def get_queryset(self):
        queryset = super().get_queryset()

        # Retrieve the branch_id from the query parameters
        branch_id = self.request.query_params.get('branch_id')

        # Filter departments based on the provided branch_id
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)

        return queryset  

    @action(detail=False, methods=['get'])
    def department_report(self, request):
        queryset = dept_master.objects.all()
        resource = DeptReportResource()
        dataset = resource.export(queryset)

        # Get the fields from the resource
        fields = [field.column_name for field in resource.get_fields()]

        # Prepare JSON response
        data = [
            {field: row[field] if field != 'dept_is_active' else (True if row[field] else False)
            for field in fields}
            for row in dataset.dict
        ]
        
        return JsonResponse(data, safe=False)
        
    
    @action(detail=False, methods=['get'])
    def deparment_excel_report(self,request):
        if request.method == 'GET':
            queryset = dept_master.objects.all()
            resource = DeptReportResource()
            dataset = resource.export(queryset)

            fields = [field.column_name for field in resource.get_fields()]
            # Create an Excel workbook
            wb = Workbook()
            ws = wb.active

            # Define default cell formats
            default_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # pink background color
            default_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Apply default styles to headings
            for col_num, field_name in enumerate(dataset.headers):
                ws.cell(row=1, column=col_num + 1, value=field_name)
                ws.cell(row=1, column=col_num + 1).fill = default_fill
                ws.cell(row=1, column=col_num + 1).alignment = default_alignment

            # Write data to the Excel file with boolean conversion
            for row_num, row in enumerate(dataset.dict, start=2):
                for col_num, field_name in enumerate(fields, start=1):
                    value = row[field_name]
                    if field_name == 'dept_is_active':
                        value = 'True' if value else 'False'
                    ws.cell(row=row_num, column=col_num, value=value)
                    ws.cell(row=row_num, column=col_num).alignment = default_alignment

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Save the workbook to an in-memory buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Prepare response
            response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="deparment_Report.xlsx"'
            return response
    @action(detail=True, methods=['POST', 'GET'])
    def policies(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            family_members = employee.policies.all()
            serializer = CompanyPolicySerializer(family_members, many=True)
            return Response(serializer.data)
@api_view(['POST'])
def save_notification_settings(request):
    selected_departments = request.data.get('departments', [])
    branch_id = request.data.get('branch_id')

    try:
        branch = brnch_mstr.objects.get(id=branch_id)
    except brnch_mstr.DoesNotExist:
        return Response({"error": "Branch not found"}, status=404)

    # Assuming you store department notification settings per branch
    branch.departments.set(selected_departments)
    branch.save()

    return Response({"status": "success"})

 #Dept BulkUpload       
class DeptBulkUploadViewSet(viewsets.ModelViewSet):
    queryset = dept_master.objects.all()
    serializer_class = DeptUploadSerializer
    

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Load data from the Excel file into a Dataset
                    dataset = Dataset()
                    dataset.load(excel_file.read(), format='xlsx')

                    # Create a resource instance
                    resource = DepartmentResource()

                    # Import data into the model using the resource
                    result = resource.import_data(dataset, dry_run=False, raise_errors=True)

                    return Response({"message": f"{result.total_rows} records created successfully"})
                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)

#DESIGNATION 
class DesignationViewSet(viewsets.ModelViewSet):
    queryset = desgntn_master.objects.all()
    serializer_class = DesgSerializer
    # authentication_classes = [SessionAuthentication,]
    permission_classes = [DesignationPermission,] 
    
    @action(detail=False, methods=['get'])
    def designation_report(self, request):
        queryset = desgntn_master.objects.all()
        resource = DesgtnReportResource()
        dataset = resource.export(queryset)

        # Get the fields from the resource
        fields = [field.column_name for field in resource.get_fields()]

        # Prepare JSON response
        data = [
            {field: row[field] if field != 'desgntn_is_active' else (True if row[field] else False)
            for field in fields}
            for row in dataset.dict
        ]
        print(data)
        return JsonResponse(data, safe=False)

    @action(detail=True, methods=['POST', 'GET'])
    def policies(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            family_members = employee.policies.all()
            serializer = CompanyPolicySerializer(family_members, many=True)
            return Response(serializer.data)   
    
    @action(detail=False, methods=['get'])
    def designation_excel_report(self,request):
        if request.method == 'GET':
            queryset = desgntn_master.objects.all()
            resource = DesgtnReportResource()
            dataset = resource.export(queryset)

            fields = [field.column_name for field in resource.get_fields()]
            # Create an Excel workbook
            wb = Workbook()
            ws = wb.active

            # Define default cell formats
            default_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # pink background color
            default_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Apply default styles to headings
            for col_num, field_name in enumerate(dataset.headers):
                ws.cell(row=1, column=col_num + 1, value=field_name)
                ws.cell(row=1, column=col_num + 1).fill = default_fill
                ws.cell(row=1, column=col_num + 1).alignment = default_alignment

            # Write data to the Excel file with boolean conversion
            for row_num, row in enumerate(dataset.dict, start=2):
                for col_num, field_name in enumerate(fields, start=1):
                    value = row[field_name]
                    if field_name == 'desgntn_is_active':
                        value = 'True' if value else 'False'
                    ws.cell(row=row_num, column=col_num, value=value)
                    ws.cell(row=row_num, column=col_num).alignment = default_alignment

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Save the workbook to an in-memory buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Prepare response
            response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="designtn_Report.xlsx"'
            return response

    

#Designation BulkUpload
class DesignationBulkUploadViewSet(viewsets.ModelViewSet):
    queryset = desgntn_master.objects.all()
    serializer_class = DesgUploadSerializer
   
    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Load data from the Excel file into a Dataset
                    dataset = Dataset()
                    dataset.load(excel_file.read(), format='xlsx')

                    # Create a resource instance
                    resource = DesignationResource()

                    # Initialize a list to capture row-wise errors
                    errors = []

                    # Validate each row before importing
                    for row_number, row in enumerate(dataset.dict, start=2):  # Start counting from row 2
                        try:
                            resource.before_import_row(row, row_number=row_number)
                        except ValueError as e:
                            # Append validation errors for this row
                            errors.append(str(e))

                    # If there are errors, return them in the response
                    if errors:
                        return Response({"validation_errors": errors}, status=400)

                    # Import data if there are no errors
                    result = resource.import_data(dataset, dry_run=False, raise_errors=True)

                    # Return success message
                    return Response({"message": f"{result.total_rows} records created successfully"})

                except Exception as e:
                    # Handle unexpected errors
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)

#CATOGARY CRUD
class CatogoryViewSet(viewsets.ModelViewSet):
    queryset = ctgry_master.objects.all()
    serializer_class = CtgrySerializer
    # authentication_classes = [SessionAuthentication,]
    permission_classes = [CategoryPermission,] 

class CategoryBulkUploadViewSet(viewsets.ModelViewSet):
    queryset = ctgry_master.objects.all()
    serializer_class = CtgryUploadSerializer
    

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Load data from the Excel file into a Dataset
                    dataset = Dataset()
                    dataset.load(excel_file.read(), format='xlsx')

                    # Create a resource instance
                    resource = CategoryResource()

                    # Import data into the model using the resource
                    result = resource.import_data(dataset, dry_run=False, raise_errors=True)

                    return Response({"message": f"{result.total_rows} records created successfully"})
                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)


class FiscalYearViewSet(viewsets.ModelViewSet):
    queryset = FiscalYear.objects.all()
    serializer_class = FiscalYearSerializer
    permission_classes = [FiscalYearPermission] 

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Create the FiscalYear instance
        self.perform_create(serializer)
        
        # Automatically create 12 FiscalPeriod instances with a gap of 30 days
        fiscal_year = serializer.instance
        start_date = fiscal_year.start_date
        for period_number in range(1, 13):
            period_start_date = start_date + timedelta(days=(period_number - 1) * 30)
            period_end_date = period_start_date + timedelta(days=29)
            FiscalPeriod.objects.create(
                fiscal_year=fiscal_year,
                period_number=period_number,
                start_date=period_start_date,
                end_date=period_end_date,
                branch=fiscal_year.branch_id
            )
        
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

class PeriodViewSet(viewsets.ModelViewSet):
    queryset = FiscalPeriod.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [FiscalYearPermission] 

class CompanyFiscalData(APIView):
    def get(self, request, company_id):
        try:
            fiscal_years = FiscalYear.objects.filter(company_id=company_id)
            fiscal_years_serializer = FiscalYearSerializer(fiscal_years, many=True)
            
            fiscal_periods = FiscalPeriod.objects.filter(company_id=company_id)
            fiscal_periods_serializer = PeriodSerializer(fiscal_periods, many=True)
            
            return Response({
                'fiscal_years': fiscal_years_serializer.data,
                'fiscal_periods': fiscal_periods_serializer.data
            })
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)
        

class permissionviewset(viewsets.ModelViewSet):
    queryset = UserTenantPermissions.objects.all()
    serializer_class = PermissionSerializer
    def get_serializer_class(self):
        return PermissionSerializer

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser:
                # Superusers see all user permissions
                return UserTenantPermissions.objects.all()
            else:
                # Regular users only see their assigned permissions
                return UserTenantPermissions.objects.filter(profile_id=user.id)
        return UserTenantPermissions.objects.none() 
    # def get_queryset(self):
    # def get_queryset(self):
    #     user = self.request.user
    #     return UserTenantPermissions.objects.filter(profile=user)  # Optional filtering

    # serializer_class = PermissionSerializer

class Groupviewset(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer

class permviewset(viewsets.ModelViewSet):
    queryset = Permission.objects.all().order_by('id')
    serializer_class=permserializer

class DocNumberingviewset(viewsets.ModelViewSet):
    queryset = DocumentNumbering.objects.all()
    serializer_class = DocumentNumberingSerializer
    permission_classes = [DocumentNumberingPermission]


from django.shortcuts import get_object_or_404
    
class FiscalYearDatesView(APIView):
    def get(self, request, fiscal_year_id):
        try:
            fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
        except FiscalYear.DoesNotExist:
            return Response({"error": "Fiscal year not found."}, status=status.HTTP_404_NOT_FOUND)

        start_date = fiscal_year.start_date
        end_date = fiscal_year.end_date

        if start_date is None or end_date is None:
            return Response({"error": "Fiscal year does not have valid start and end dates."}, status=status.HTTP_400_BAD_REQUEST)

        date_list = []
        current_date = start_date
        while current_date <= end_date:
            date_list.append(current_date)
            current_date += timedelta(days=1)

        return Response({
            "fiscal_year": fiscal_year.id,
            "branch": fiscal_year.branch_id.id,
            "dates": date_list
        })
    
class FiscalPeriodDatesView(APIView):
    def get(self, request, fiscal_year_id, period_number):
        try:
            fiscal_period = FiscalPeriod.objects.get(fiscal_year_id=fiscal_year_id, period_number=period_number)
        except FiscalPeriod.DoesNotExist:
            return Response({"error": "Fiscal period not found."}, status=404)

        start_date = fiscal_period.start_date
        end_date = fiscal_period.end_date

        date_list = []
        current_date = start_date
        while current_date <= end_date:
            date_list.append(current_date)
            current_date += timedelta(days=1)

        return Response({
            "fiscal_year": fiscal_period.fiscal_year.id,
            "period_number": fiscal_period.period_number,
            "branch": fiscal_period.branch.id,
            "dates": date_list
        })
    
class CompanyPolicyViewSet(viewsets.ModelViewSet):
    queryset = CompanyPolicy.objects.all()
    serializer_class = CompanyPolicySerializer
    permission_classes = [CompanyPolicyPermission]
    
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser:
                return self.queryset

            # Check if user is listed in `specific_users`
            if self.queryset.filter(specific_users=user).exists():
                return self.queryset.filter(specific_users=user)

            # Get the employee's branch, department, and category if the user is an ESS user
            if user.is_ess:
                emp_branch = user.emp_master.emp_branch_id
                emp_dept = user.emp_master.emp_dept_id
                emp_category = user.emp_master.emp_ctgry_id

                # Filter policies by matching the branch, department, and category
                return self.queryset.filter(
                    branch=emp_branch,
                    department=emp_dept,
                    category=emp_category
                )
        
        return CompanyPolicy.objects.none()
    # def get_queryset(self):
    #     user = self.request.user
    #     if user.is_authenticated:
    #         if user.is_superuser:
    #             return self.queryset
    #         # Get the employee's branch, department, and category
    #         emp_branch = user.emp_master.emp_branch_id
    #         emp_dept = user.emp_master.emp_dept_id
    #         emp_category = user.emp_master.emp_ctgry_id

    #         # Filter policies by matching the branch, department, and category
    #         return self.queryset.filter(
    #             branch=emp_branch,
    #             department=emp_dept,
    #             category=emp_category
    #         )
    #     return CompanyPolicy.objects.none()
    
    @action(detail=True, methods=['get'], url_path='download', url_name='download_policy')
    def download_policy(self, request, pk=None):
        policy = self.get_object()

        if not policy.policy_file:
            return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)

        file_extension = os.path.splitext(policy.policy_file.name)[1].lower()

        if file_extension == '.docx':
            return self.convert_docx_to_pdf(policy)

        elif file_extension == '.xlsx':
            return self.convert_xlsx_to_pdf(policy)
        elif file_extension == '.txt':
            return self.convert_txt_to_pdf(policy)

        elif file_extension == '.pdf':
            try:
                response = FileResponse(policy.policy_file.open(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{policy.title}.pdf"'
                return response
            except FileNotFoundError:
                raise Http404("File not found")

        else:
            return Response({'error': 'Unsupported file format'}, status=status.HTTP_400_BAD_REQUEST)

    def convert_docx_to_pdf(self, policy):
        """
        Convert .docx content to a readable paragraph format in a PDF.
        """
        doc = Document(policy.policy_file)
        buffer = BytesIO()
        pdf_canvas = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y_position = height - 50  # Starting position for text
        
        # Read each paragraph and add it as text to the PDF
        for para in doc.paragraphs:
            if y_position < 50:  # Check if page needs to be advanced
                pdf_canvas.showPage()
                y_position = height - 50
            text = para.text
            pdf_canvas.drawString(50, y_position, text)
            y_position -= 20  # Move down for the next line

        pdf_canvas.save()

        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f'{policy.title}.pdf')

    def convert_xlsx_to_pdf(self, policy):
        """
        Convert .xlsx content to a paragraph-like format in a PDF.
        """
        policy.policy_file.open()
        workbook = load_workbook(policy.policy_file)
        sheet = workbook.active

        buffer = BytesIO()
        pdf_canvas = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y_position = height - 50  # Starting position for text

        # Iterate over rows and columns to format as paragraph-like text
        for row in sheet.iter_rows(values_only=True):
            row_text = ', '.join([str(cell) for cell in row if cell is not None])  # Join all cells with a comma
            if y_position < 50:  # Check if page needs to be advanced
                pdf_canvas.showPage()
                y_position = height - 50

            pdf_canvas.drawString(50, y_position, row_text)
            y_position -= 20  # Move down for the next line

        pdf_canvas.save()

        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f'{policy.title}.pdf')

    def convert_txt_to_pdf(self, policy):
        """
        Convert .txt content to a readable paragraph format in a PDF.
        """
        policy.policy_file.open()
        buffer = BytesIO()
        pdf_canvas = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y_position = height - 50  # Starting position for text

        # Read the .txt file line by line and write it to the PDF
        for line in policy.policy_file:
            if y_position < 50:  # Check if page needs to be advanced
                pdf_canvas.showPage()
                y_position = height - 50

            # Decode if it's a binary file (bytes) to text
            text = line.decode('utf-8').strip()
            pdf_canvas.drawString(50, y_position, text)
            y_position -= 20  # Move down for the next line

        pdf_canvas.save()

        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f'{policy.title}.pdf')
    
def list_data_in_schema(request):
    schema_name = request.GET.get('schema')  # Use explicit schema name if tenant is unavailable
    if not schema_name:
        return JsonResponse({"error": "Schema name is required"}, status=400)

    try:
        with schema_context(schema_name):
            branches = brnch_mstr.objects.all()
            departments = dept_master.objects.all()
            designations = desgntn_master.objects.all()
            categories = ctgry_master.objects.all()

            data = {
                'branches': [{'id': branch.id, 'name': branch.branch_name} for branch in branches],
                'departments': [{'id': dept.id, 'name': dept.dept_name} for dept in departments],
                'designations': [{'id': desig.id, 'name': desig.desgntn_job_title} for desig in designations],
                'categories': [{'id': cat.id, 'name': cat.ctgry_title} for cat in categories],
            }

            return JsonResponse(data)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

class AnnouncementViewSet(viewsets.ModelViewSet):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    # permission_classes = [permissions.IsAuthenticated]

    # def get_queryset(self):
    #     user = self.request.user
    #     employee = getattr(user, 'employee', None)

    #     if not employee:
    #         return Announcement.objects.none()

    #     return Announcement.objects.filter(
    #         Q(specific_employees=employee) | Q(branches=employee.branch),
    #         Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
    #     ).distinct().order_by('-is_sticky', '-created_at')

    def perform_create(self, serializer):
        instance = serializer.save()
        self.send_announcement_emails(instance)

    def send_announcement_emails(self, announcement):
        

        if not announcement.send_email:
            return

        config = EmailConfiguration.objects.filter(is_active=True).first()
        if not config or not config.email_host_user or not config.email_host_password:
            return

        recipients = set()
        recipients.update(emp.emp_personal_email for emp in announcement.specific_employees.all() if emp.emp_personal_email)
        for branch in announcement.branches.all():
            branch_employees = emp_master.objects.filter(emp_branch_id=branch)
            recipients.update(emp.emp_personal_email for emp in branch_employees if emp.emp_personal_email)

        subject = f"{'[Sticky] ' if announcement.is_sticky else ''}New Announcement: {announcement.title}"
        # body = announcement.message
        body = f"""
        Dear {announcement.message},

        {announcement.message}

        {'[Sticky]' if announcement.is_sticky else ''}
        {'[Expires on {}]'.format(announcement.expires_at) if announcement.expires_at else ''}

        Regards,
        HR Team
        """
        email = EmailMessage(
            subject,
            body,
            config.email_host_user,
            list(recipients),
        )

        if announcement.attachment:
            email.attach_file(announcement.attachment.path)

        email.send(fail_silently=True)
    # Mark announcement as viewed
    @action(detail=True, methods=['post'])
    def mark_as_viewed(self, request, pk=None):
        announcement = self.get_object()
        emp_id = request.data.get('employee_id')
        try:
            employee = emp_master.objects.get(id=emp_id)
        except emp_master.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        view_obj, created = AnnouncementView.objects.get_or_create(announcement=announcement, employee=employee)
        return Response({'status': 'viewed', 'created': created})
class AnnouncementCommentViewSet(viewsets.ModelViewSet):
    queryset = AnnouncementComment.objects.all()
    serializer_class = AnnouncementCommentSerializer

    def create(self, request, *args, **kwargs):
        announcement_id = request.data.get('announcement')
        employee_id = request.data.get('employee')
        try:
            announcement = Announcement.objects.get(id=announcement_id)
            if not announcement.allow_comments:
                return Response({'error': 'Comments are disabled for this announcement.'}, status=400)
        except Announcement.DoesNotExist:
            return Response({'error': 'Announcement not found'}, status=404)

        return super().create(request, *args, **kwargs)

class AssetTypeViewSet(viewsets.ModelViewSet):
    queryset = AssetType.objects.all()
    serializer_class = AssetTypeSerializer
  
class AssetMasterViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer

class AssetRequestViewSet(viewsets.ModelViewSet):
    queryset = AssetRequest.objects.all()
    serializer_class = AssetRequestSerializer
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        asset_request = self.get_object()
        
        # Check if an asset is already assigned to the request
        if not asset_request.requested_asset:
            # Try to assign an available asset of the requested type
            available_asset = Asset.objects.filter(
                asset_type=asset_request.asset_type,
                status="available"
            ).first()
            
            if not available_asset:
                return Response(
                    {"error": "No available asset of the requested type."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Assign the asset to the request
            asset_request.requested_asset = available_asset
            asset_request.save()

        # Proceed with approval
        asset = asset_request.requested_asset

        if asset.status != "available":
            return Response(
                {"error": f"Asset '{asset.name}' is not available."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Approve the request and allocate the asset
        with transaction.atomic():  # Ensures all database updates happen together
            # Update the request status
            asset_request.status = "approved"
            asset_request.save()

            # Update the asset status
            asset.status = "assigned"
            asset.save()

            # Create the AssetAllocation entry
            AssetAllocation.objects.create(
                asset=asset,
                employee=asset_request.employee,
                assigned_date=timezone.now().date(),
            )

        return Response(
            {
                "status": "approved",
                "message": f"Asset '{asset.name}' has been allocated to {asset_request.employee}."
            },
            status=status.HTTP_200_OK
        )
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        asset_request = self.get_object()
        asset_request.status = "rejected"
        asset_request.save()
        return Response({"status": "rejected"}, status=status.HTTP_200_OK)

class AssetAllocationViewSet(viewsets.ModelViewSet):
    queryset = AssetAllocation.objects.all()
    serializer_class = AssetAllocationSerializer
    @action(detail=True, methods=['post'])
    def return_asset(self, request, pk=None):
        allocation = self.get_object()
        condition = request.data.get('return_condition')
        returned_date = request.data.get('returned_date')  # Optional
        
        if not condition:
            return Response(
                {"error": "Return condition is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            allocation.return_asset(condition, returned_date)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"message": "Asset returned successfully."},
            status=status.HTTP_200_OK
        )
    
class Asset_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = AssetCustomFieldValue.objects.all()
    serializer_class = AssetCustomFieldValueSerializer

class AssetCustomFieldViewSet(viewsets.ModelViewSet):
    queryset = AssetCustomField.objects.all()
    serializer_class = AssetCustomFieldSerializer


class AssetReportViewset(viewsets.ModelViewSet):
    queryset = AssetReport.objects.all()
    serializer_class = AssetReportSerializer
    
    def __init__(self, *args, **kwargs):
        super(AssetReportViewset, self).__init__(*args, **kwargs)
        self.ensure_standard_report_exists()

    def get_available_fields(self):
        excluded_fields = {'id'}
        display_names = {
            "asset_type": "Asset Type",
            "name": "Name",
            "model": "Model",
            "purchase_date": "Purchase Date",
            "status": "Status",
            "condition": "Condition",
        }
        
        asset_fields = [field.name for field in Asset._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        custom_fields = list(AssetCustomField.objects.values_list('name', flat=True))        
        available_fields = {field: display_names.get(field, field) for field in asset_fields + custom_fields} 
        return available_fields

    @action(detail=False, methods=['get'])
    def select_employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
        

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def emp_select_report(self, request, *args, **kwargs):
        # if not request.user.is_superuser:
        #     return Response({"error": "You do not have permission to access this resource."}, status=status.HTTP_403_FORBIDDEN)
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'reports')  # Default to 'report' if 'file_name' is not provided
                fields_to_include = request.POST.getlist('fields', [])
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())

            assets = Asset.objects.all()

            report_data = self.generate_report_data(fields_to_include, assets)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')  # Use 'file_name' provided by the user

            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format

            AssetReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({
                'status': 'success',
                'file_path': file_path,
                'selected_fields_data': fields_to_include,
                
            })

        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    def ensure_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if AssetReport.objects.filter(file_name='std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'std_report'
            fields_to_include = self.get_available_fields().keys()
            asset = Asset.objects.all()

            report_data = self.generate_report_data(fields_to_include, asset)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            AssetReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = AssetReport.objects.get(file_name='std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except AssetReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include, assets):
        asset_fields = [field.name for field in Asset._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        report_data = []

        for asset in assets:
            # Get custom fields specific to the asset type
            asset_custom_fields = AssetCustomField.objects.filter(asset_type=asset.asset_type).values_list('name', flat=True)
            employee_data = {}

            for field in fields_to_include:
                if field in asset_fields:
                    # Handle standard asset fields
                    value = getattr(asset, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()  # Convert date to ISO format string
                elif field in asset_custom_fields:
                    # Handle custom fields specific to the asset type
                    custom_field_value = AssetCustomFieldValue.objects.filter(
                        asset=asset,
                        custom_field__name=field
                    ).first()
                    value = custom_field_value.field_value if custom_field_value else 'N/A'
                else:
                    # Field not found
                    value = 'N/A'

                employee_data[field] = value
            report_data.append(employee_data)

        return report_data
    
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])  # Get selected fields from session
        print("selected fields:",selected_fields)
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })
    
    @csrf_exempt
    @action(detail=False, methods=['post'])
    def generate_employee_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.POST.get('report_id')
        available_fields = self.get_available_fields()
       
        # Save selected fields to session
        request.session['selected_fields'] = selected_fields
        print("select fields",selected_fields)
        # Fetch report data based on report_id
        try:
            report = AssetReport.objects.get(id=report_id)
            report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)  # Assuming report_data is a FileField
            with open(report_file_path, 'r') as file:
                report_content = json.load(file)  # Load content of the report file as JSON
        except AssetReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'})
        print("reportcontnt",report_content)
        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if report_content:
                selected_fields = list(report_content[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report

        # Fetch employees data from emp_master
        employees = Asset.objects.all()

        # Get unique values for selected_fields
        unique_values = self.get_unique_values_for_fields(employees, selected_fields, report_content)
        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': report_content,  # Pass report_content to the frontend
            'unique_values': processed_unique_values,
        })

       

    def get_unique_values_for_fields(self, employees, selected_fields, report_content):
        unique_values = {field: set() for field in selected_fields}

        # Extract data from the JSON content
        for record in report_content:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Fetch additional data from Emp_CustomField if necessary
        for field in selected_fields:
            if field not in unique_values:
                continue
            for employee in employees:
                if not hasattr(employee, field):
                    custom_field_value = AssetCustomField.objects.filter(Asset=employee, field_name=field).first()
                    if custom_field_value:
                        unique_values[field].add(custom_field_value.field_value)

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
    
    
    @csrf_exempt
    @action(detail=False, methods=['post'])
    def filter_existing_report(self, request, *args, **kwargs):
        report_id = request.data.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        try:
            report_instance = AssetReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except (AssetReport.DoesNotExist, json.JSONDecodeError) as e:
            return HttpResponse(f'Report not found or invalid JSON format: {str(e)}', status=404)

        selected_fields = [key for key in request.data.keys() if key != 'report_id']
        filter_criteria = {}

        for field in selected_fields:
            values = [val.strip() for val in request.data.getlist(field) if val.strip()]
            if values:
                filter_criteria[field] = values

        field_names = {
            "Asset Type":"asset_type",
            "Name":"name",
            "Model":"model" ,
            "Purchase Date":"purchase_date",
            "Status":"status",
            "Condition":"condition",
            # Add other field mappings as per your needs
        }

        filtered_data = [row for row in report_data if self.match_filter_criteria(row, filter_criteria, field_names)]
        print("filtered data",filtered_data)
        # Save filtered data to session for Excel generation
        request.session['filtered_data'] = filtered_data
        request.session.modified = True
        display_named = self.get_available_fields()

        return JsonResponse({
        'filtered_data': filtered_data,
        'report_id': report_id,
    })
        

    def match_filter_criteria(self, row_data, filter_criteria, field_names):
        for column_heading, field_name in field_names.items():
            if field_name in filter_criteria:
                values = filter_criteria[field_name]
                row_value = row_data.get(field_name)
                if row_value is None or row_value.strip() not in values:
                    return False
        for custom_field_name in filter_criteria.keys():
            if custom_field_name not in field_names.values():
                custom_field_values = filter_criteria[custom_field_name]
                custom_field_value = row_data.get(custom_field_name, '').strip().lower()
                if custom_field_value and custom_field_value not in [val.lower() for val in custom_field_values]:
                    return False
        return True
class AssetTransactionReportViewset(viewsets.ModelViewSet):
    queryset = AssetTransactionReport.objects.all()
    serializer_class = AssetTransactionReportSerializer

    def get_available_fields(self):
        excluded_fields = {'id'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id','emp_branch_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            'emp_branch_id':"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "asset":"Asset",
            "assigned_date":"Assigned Date",
            "returned_date":"returned_date",
            "return_condition":"Return Condition"


           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        asset_fields = [field.name for field in AssetAllocation._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + asset_fields}
        return available_fields
    @action(detail=False, methods=['get'])
    def select_asset_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
    
    @action(detail=False, methods=['post'])
    def generate_general_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            generalreport = AssetAllocation.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,generalreport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            AssetTransactionReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def general_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if AssetTransactionReport.objects.filter(file_name='generalrequest_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'generalrequest_std_report'
            fields_to_include = self.get_available_fields().keys()
            generalreport = AssetAllocation.objects.all()

            report_data = self.generate_report_data(fields_to_include, generalreport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            AssetTransactionReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )
        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = AssetTransactionReport.objects.get(file_name='generalrequest_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except AssetTransactionReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include,generalreport):
        column_headings = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "branch": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "asset":"Asset",
            "assigned_date":"Assigned Date",
            "returned_date":"returned_date",
            "return_condition":"Return Condition"
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        general_request_fields = [field.name for field in AssetAllocation._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = []
        for document in generalreport:
            general_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in general_request_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                general_data[field] = value
            report_data.append(general_data)
        return report_data            

    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        }) 
      
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = AssetTransactionReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except AssetTransactionReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'created_at_date' in row and row['assigned_date'] and
            start_date <= datetime.fromisoformat(row['assigned_date']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
    
    @action(detail=False, methods=['post'])
    def generate_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Fetch previously filtered date data from the `apply_date_filter` method
        date_filtered_data = getattr(self, 'date_filtered_data', [])
        print("previous",date_filtered_data)

        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            report_instance = get_object_or_404(AssetTransactionReport, id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
            date_filtered_data = report_data

        # Default to all fields if no specific fields selected
        if not selected_fields and date_filtered_data:
            selected_fields = list(date_filtered_data[0].keys())

        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {
            field: {'values': values}
            for field, values in unique_values.items()
        }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,
            'unique_values': processed_unique_values,
            # 'column_headings': column_headings
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
    @action(detail=False, methods=['post'])
    def general_filter_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)
        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]
        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })