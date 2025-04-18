from django.shortcuts import render
from django.conf import settings
from datetime import date
import logging
from django.utils import timezone
from openpyxl import load_workbook
from .models import (emp_family,Emp_Documents,EmpJobHistory,EmpLeaveRequest,EmpQualification,GeneralRequest,RequestType,
                     emp_master,notification,EmpFamily_CustomField,EmpJobHistory_CustomField,
                     EmpQualification_CustomField,EmpDocuments_CustomField,LanguageSkill,MarketingSkill,ProgrammingLanguageSkill,Emp_CustomField,Report,Doc_Report,GeneralRequest,RequestType,GeneralRequestReport,EmployeeLangSkill,EmployeeProgramSkill,
                     EmployeeMarketingSkill,Approval,ApprovalLevel,RequestNotification,Emp_CustomFieldValue,
                     EmailTemplate,EmailConfiguration,SelectedEmpNotify,NotificationSettings,DocExpEmailTemplate,CommonWorkflow,Doc_CustomFieldValue,EmployeeBankDetail,Fam_CustomFieldValue,Qualification_CustomFieldValue,JobHistory_CustomFieldValue
                     )
from .serializer import (Emp_qf_Serializer,EmpFamSerializer,EmpSerializer,NotificationSerializer,RequestTypeSerializer,
                         EmpJobHistorySerializer,EmpLeaveRequestSerializer,DocumentSerializer,GeneralRequestSerializer,
                         GeneralReportSerializer,EmpMarketSkillSerializer,EmployeeReportSerializer,EmpBulkUploadSerializer,CustomFieldSerializer,
                         EmpFam_CustomFieldSerializer,EmpJobHistory_Udf_Serializer,Emp_qf_udf_Serializer,EmpDocuments_Udf_Serializer,
                         DocBulkuploadSerializer,DocumentReportSerializer,EmpPrgrmSkillSerializer,EmpLangSkillSerializer,ApprovalSerializer,ApprovalLevelSerializer,
                         ReqNotifySerializer,Emp_CustomFieldValueSerializer,EmailTemplateSerializer,EmployeeFilterSerializer,EmailConfigurationSerializer,SelectedEmpNotifySerializer,
                         NotificationSettingsSerializer,DocExpEmailTemplateSerializer,CommonWorkflowSerializer,DOC_CustomFieldValueSerializer,EmpBankDetailsSerializer,EmpBankBulkuploadSerializer,EmplistSerializer,Fam_CustomFieldValueSerializer,
                         Qualification_CustomFieldValueSerializer,JobHistory_CustomFieldValueSerializer)

from .resource import EmployeeResource,DocumentResource,EmpCustomFieldValueResource,EmpDocumentCustomFieldValueResource,EmpBankDetailsResource, MarketingSkillResource,ProLangSkillResource
from .permissions import (IsSuperUserOrHasGeneralRequestPermission,IsSuperUserOrInSameBranch,EmpCustomFieldPermission,EmpCustomFieldValuePermission,
                        EmpFamilyCustomFieldPermission,EmpJobHistoryCustomFieldPermission,EmpQualificationCustomFieldPermission,ReportPermission,DocReportPermission,GeneralRequestReportPermission,
                        EmployeeMarketingSkillPermission,EmployeeProgramSkillPermission,EmployeeLangSkillPermission,NotificationPermission,ApprovalLevelPermission,EmployeeMarketingSkillPermission,RequestTypePermission,EmployeeListPermission)
from django.core.exceptions import ValidationError
from rest_framework.decorators import action
from phonenumber_field.modelfields import PhoneNumberField
from rest_framework import viewsets,filters
from tablib import Dataset
from io import BytesIO
import os,json
from django.db.models import Field
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Alignment,Font,NamedStyle
from rest_framework import status,generics,viewsets,permissions
from .permissions import EmployeePermission
from datetime import datetime, timedelta
from OrganisationManager.models import DocumentNumbering
from OrganisationManager.serializer import DocumentNumberingSerializer
# from rest_framework.authentication import SessionAuthentication,TokenAuthentication
from rest_framework.permissions import IsAuthenticated,AllowAny,IsAuthenticatedOrReadOnly,IsAdminUser
# from django. utils. timezone import timedelta
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd,openpyxl
from django.views.decorators.csrf import csrf_exempt
from rest_framework.authentication import BaseAuthentication
from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect
from django.contrib.contenttypes.models import ContentType
from rest_framework import views, status
from django.forms.models import model_to_dict
from .tasks import send_document_expiry_notifications_for_all_tenants
from django.core.cache import cache
import redis
import json
from calendars .serializer import AttendanceSerializer
from rest_framework.exceptions import NotFound
from calendars .serializer import EmployeeLeaveBalanceSerializer,LeaveTypeSerializer
from calendars .models import leave_type, employee_leave_request
from django.db.models import Q

r = redis.StrictRedis(host='localhost', port=6379, db=0)

class CustomAuthentication(BaseAuthentication):
    def authenticate(self, request):
        user = request.user
        if user.is_authenticated:
            companies = user.companies.all()
            if companies:
                # If user has associated companies, return the user and None
                return user, None
            else:
                # If user has no associated companies, raise PermissionDenied
                raise PermissionDenied("You do not have access to any schemas.")
        return None

#EMPLOYEE CRUD
class EmpViewSet(viewsets.ModelViewSet):
    queryset = emp_master.objects.all()
    serializer_class = EmpSerializer
    permission_classes = [EmployeePermission]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if hasattr(user, 'is_ess') and user.is_ess:  # If user is an ESS, they can only access their own employee information
                return emp_master.objects.filter(users=user)
            else:
                return emp_master.objects.all()  # Other users can access all employee information
        return emp_master.objects.none()
    @action(detail=True, methods=['POST', 'GET'])
    def emp_family(self, request, pk=None):
        employee = self.get_object()

        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpFamSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_family.all()
            serializer = EmpFamSerializer(family_members, many=True)
            return Response(serializer.data)
    
    @action(detail=True, methods=['POST', 'GET'])
    def emp_qualification(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
        # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = Emp_qf_Serializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_qualification.all()
            serializer = Emp_qf_Serializer(family_members, many=True)
            return Response(serializer.data)

    @action(detail=True, methods=['POST', 'GET'])
    def emp_job_history(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpJobHistorySerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_job_history.all()
            serializer = EmpJobHistorySerializer(family_members, many=True)
            return Response(serializer.data)
    



    @action(detail=True, methods=['POST', 'GET','DELETE'])
    def emp_documents(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = DocumentSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_documents.all()
            serializer = DocumentSerializer(family_members, many=True)
            return Response(serializer.data)
    
    # @action(detail=True, methods=['POST', 'GET'])
    @action(detail=True, methods=['POST', 'GET', 'DELETE'])
    def emp_market_skills(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpMarketSkillSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_market_skills.all()
            serializer = EmpMarketSkillSerializer(family_members, many=True)
            return Response(serializer.data)
            
        elif request.method == 'DELETE':
            # Delete all emp_market_skills related to the employee
            skills = employee.emp_market_skills.all()
            skills_count = skills.count()
            skills.delete()
            return Response({"detail": f"{skills_count} skills deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


    @action(detail=True, methods=['POST', 'GET','DELETE'])
    def emp_programlangskill(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
        # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpPrgrmSkillSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_prgrm_skills.all()
            serializer = EmpPrgrmSkillSerializer(family_members, many=True)
            return Response(serializer.data)

        elif request.method == 'DELETE':
            # Delete all emp_market_skills related to the employee
            skills = employee.emp_prgrm_skills.all()
            skills_count = skills.count()
            skills.delete()
            return Response({"detail": f"{skills_count} skills deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


    @action(detail=True, methods=['POST', 'GET','DELETE'])
    def emp_languageskill(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
        # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpLangSkillSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_lang_skills.all()
            serializer = EmpLangSkillSerializer(family_members, many=True)
            return Response(serializer.data)
        
        elif request.method == 'DELETE':
            # Delete all emp_market_skills related to the employee
            skills = employee.emp_lang_skills.all()
            skills_count = skills.count()
            skills.delete()
            return Response({"detail": f"{skills_count} skills deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    
    
    @action(detail=True, methods=['get'])
    def attendance(self, request, pk=None):
        user = self.get_object()
        attendance = user.get_attendance()
        serializer = AttendanceSerializer(attendance, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def leave_balance(self, request, pk=None):
        user = self.get_object()

        # Get leave balance
        leave_balance = user.get_leave_balance()

        # Get pending leave requests for the employee
        pending_leave_types = employee_leave_request.objects.filter(
            employee=user, status="pending"
        ).values_list("leave_type", flat=True)  # Extract leave_type IDs

        # Exclude leave types that are in pending requests
        available_leave_types = leave_type.objects.exclude(id__in=pending_leave_types)

        # Serialize data
        leave_balance_serializer = EmployeeLeaveBalanceSerializer(leave_balance, many=True)
        leave_type_serializer = LeaveTypeSerializer(available_leave_types, many=True)

        return Response({
            "leave_balance": leave_balance_serializer.data,
            "available_leave_types": leave_type_serializer.data,
        }, status=status.HTTP_200_OK)
    
     

    @action(detail=False, methods=['get'])
    def filter_empty_user_non_ess(self, request):
        filtered_employees = self.queryset.filter(users__isnull=True, is_ess=False)
        serializer = EmployeeFilterSerializer(filtered_employees, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export_employee_data(self, request):
        excluded_fields = {'id', 'is_ess', 'created_at', 'created_by', 'updated_at', 'updated_by', 'emp_profile_pic'}
        display_names = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category"
        }

        # Fetch all employees
        employees = emp_master.objects.all()

        # Fetch all distinct field names from Emp_CustomField
        custom_fields = Emp_CustomField.objects.values_list('emp_custom_field', flat=True).distinct()

        # Prepare headers combining emp_master fields and custom_fields
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        headers = emp_master_fields + list(custom_fields)

        # Create an Excel workbook
        wb = Workbook()
        ws = wb.active

        # Define default cell formats
        default_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # pink background color
        default_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write headers with display names
        for col_num, header in enumerate(headers, start=1):
            display_name = display_names.get(header, header.capitalize())
            ws.cell(row=1, column=col_num, value=display_name)  # Capitalize header names
            ws.cell(row=1, column=col_num).fill = default_fill
            ws.cell(row=1, column=col_num).alignment = default_alignment

        # Write employee data and custom fields
        for row_num, employee in enumerate(employees, start=2):
            for col_num, header in enumerate(headers, start=1):
                if header in custom_fields:
                    # Fetch custom field value
                    custom_field_value = Emp_CustomFieldValue.objects.filter(emp_master=employee, emp_custom_field=header).first()
                    value = custom_field_value.field_value if custom_field_value else ''
                else:
                    value = getattr(employee, header, '')

                ws.cell(row=row_num, column=col_num, value=str(value))
                ws.cell(row=row_num, column=col_num).alignment = default_alignment

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Save the workbook to an in-memory buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Prepare response
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Empployee_data.xlsx"'
        return response



class ReportViewset(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = EmployeeReportSerializer
    permission_classes = [ReportPermission]
    # permission_classes = [IsSuperUserOrInSameBranch]

     
    # def get_queryset(self):
    #     user = self.request.user
    #     print ("user",user)
    #     # If superuser, return all reports
    #     if user.is_superuser:
    #         return Report.objects.all()
       
    #    # Filter reports based on user's branch
    #     if user.is_authenticated:
    #         print("authenticated")
    #         if hasattr(user, 'branches'):
            
    #             print("all")  
    #             user_branch_id = user.branches
    #             print("branch",user_branch_id)
    #             return Report.objects.filter(branch_id=user_branch_id)
            
    #     # If user does not have branch_id (should not happen with proper user model setup), return empty queryset or handle as needed
    #     return Report.objects.none()
   

    def __init__(self, *args, **kwargs):
        super(ReportViewset, self).__init__(*args, **kwargs)
        self.ensure_standard_report_exists()

    def get_available_fields(self):
        excluded_fields = {'id', 'is_ess','created_at', 'created_by', 'updated_at', 'updated_by', 'emp_profile_pic'}
        display_names = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category"
        }
        
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        emp_custom_fields = list(Emp_CustomField.objects.values_list('emp_custom_field', flat=True))        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + emp_custom_fields} 
        return available_fields

    @action(detail=False, methods=['get'])
    def select_employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
        

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def emp_select_report(self, request, *args, **kwargs):
        # if not request.user.is_superuser:
        #     return Response({"error": "You do not have permission to access this resource."}, status=status.HTTP_403_FORBIDDEN)
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'reports')  # Default to 'report' if 'file_name' is not provided
                fields_to_include = request.POST.getlist('fields', [])
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())

            employees = emp_master.objects.all()

            report_data = self.generate_report_data(fields_to_include, employees)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')  # Use 'file_name' provided by the user

            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format

            Report.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({
                'status': 'success',
                'file_path': file_path,
                'selected_fields_data': fields_to_include,
                
            })

        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    def ensure_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if Report.objects.filter(file_name='std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'std_report'
            fields_to_include = self.get_available_fields().keys()
            employees = emp_master.objects.all()

            report_data = self.generate_report_data(fields_to_include, employees)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            Report.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = Report.objects.get(file_name='std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except Report.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    

    def generate_report_data(self, fields_to_include, employees):
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        custom_fields = list(Emp_CustomFieldValue.objects.filter(emp_master__in=employees).values_list('emp_custom_field', flat=True).distinct())

        report_data = []
        for employee in employees:
            employee_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()  # Convert date to ISO format string
                elif field in custom_fields:
                    # Fetch the custom field value directly from Emp_CustomFieldValue
                    custom_field_value = Emp_CustomFieldValue.objects.filter(
                        emp_master=employee, 
                        emp_custom_field=field
                    ).first()
                    value = custom_field_value.field_value if custom_field_value else 'N/A'
                else:
                    value = 'N/A'
                employee_data[field] = value
            report_data.append(employee_data)

        return report_data

    
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])  # Get selected fields from session
        print("selected fields:",selected_fields)
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })   

        

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def generate_employee_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.POST.get('report_id')
        available_fields = self.get_available_fields()
       
        # Save selected fields to session
        request.session['selected_fields'] = selected_fields
        print("select fields",selected_fields)
        # Fetch report data based on report_id
        try:
            report = Report.objects.get(id=report_id)
            report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)  # Assuming report_data is a FileField
            with open(report_file_path, 'r') as file:
                report_content = json.load(file)  # Load content of the report file as JSON
        except Report.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'})
        print("reportcontnt",report_content)
        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if report_content:
                selected_fields = list(report_content[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report

        # Fetch employees data from emp_master
        employees = emp_master.objects.all()

        # Get unique values for selected_fields
        unique_values = self.get_unique_values_for_fields(employees, selected_fields, report_content)
        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': report_content,  # Pass report_content to the frontend
            'unique_values': processed_unique_values,
        })

       

    def get_unique_values_for_fields(self, employees, selected_fields, report_content):
        unique_values = {field: set() for field in selected_fields}

        # Extract data from the JSON content
        for record in report_content:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Fetch additional data from Emp_CustomField if necessary
        for field in selected_fields:
            if field not in unique_values:
                continue
            for employee in employees:
                if not hasattr(employee, field):
                    custom_field_value = Emp_CustomField.objects.filter(emp_master=employee, field_name=field).first()
                    if custom_field_value:
                        unique_values[field].add(custom_field_value.field_value)

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
    
    
    @csrf_exempt
    @action(detail=False, methods=['post'])
    def filter_existing_report(self, request, *args, **kwargs):
        report_id = request.data.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        try:
            report_instance = Report.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except (Report.DoesNotExist, json.JSONDecodeError) as e:
            return HttpResponse(f'Report not found or invalid JSON format: {str(e)}', status=404)

        selected_fields = [key for key in request.data.keys() if key != 'report_id']
        filter_criteria = {}

        for field in selected_fields:
            values = [val.strip() for val in request.data.getlist(field) if val.strip()]
            if values:
                filter_criteria[field] = values

        field_names = {
            "Employee Code": "emp_code",
            "First Name": "emp_first_name",
            "Last Name": "emp_last_name",
            "Gender": "emp_gender",
            "Date of Birth": "emp_date_of_birth",
            "Email": "emp_personal_email",
            "Mobile Number": "emp_mobile_number_1",
            "Mobile Number2": "emp_mobile_number_2",
            "Country": "emp_country_id",
            "State": "emp_state_id",
            "City": "emp_city",
            "Permanent Address": "emp_permenent_address",
            "Present Address": "emp_present_address",
            "Status": "emp_status",
            "Hired Date": "emp_hired_date",
            "Active Date": "emp_active_date",
            "Religion": "emp_relegion",
            "Blood Group": "emp_blood_group",
            "Nationality": "emp_nationality_id",
            "Marital Status": "emp_marital_status",
            "Father Name": "emp_father_name",
            "Mother Name": "emp_mother_name",
            "Posting Location": "emp_posting_location",
            "Active": "is_active",
            "OT Applicable": "epm_ot_applicable",
            "Company": "emp_company_id",
            "Branch": "emp_branch_id",
            "Department": "emp_dept_id",
            "Designation": "emp_desgntn_id",
            "Category": "emp_ctgry_id",
            # Add other field mappings as per your needs
        }

        filtered_data = [row for row in report_data if self.match_filter_criteria(row, filter_criteria, field_names)]
        print("filtered data",filtered_data)
        # Save filtered data to session for Excel generation
        request.session['filtered_data'] = filtered_data
        request.session.modified = True
        display_named = self.get_available_fields()

        return JsonResponse({
        'filtered_data': filtered_data,
        'report_id': report_id,
    })
        

    def match_filter_criteria(self, row_data, filter_criteria, field_names):
        for column_heading, field_name in field_names.items():
            if field_name in filter_criteria:
                values = filter_criteria[field_name]
                row_value = row_data.get(field_name)
                if row_value is None or row_value.strip() not in values:
                    return False
        for custom_field_name in filter_criteria.keys():
            if custom_field_name not in field_names.values():
                custom_field_values = filter_criteria[custom_field_name]
                custom_field_value = row_data.get(custom_field_name, '').strip().lower()
                if custom_field_value and custom_field_value not in [val.lower() for val in custom_field_values]:
                    return False
        return True

    

    
    @action(detail=False, methods=['get'])
    def generate_excel(self, request, *args, **kwargs):
        report_id = request.GET.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        filtered_data = request.session.get('filtered_data')
        if not filtered_data:
            return HttpResponse('No filtered data available', status=400)

        # Mapping of internal field names to display names
        field_names_mapping = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
        }

        try:
            report_instance = Report.objects.get(id=int(report_id))
        except (Report.DoesNotExist, ValueError):
            return HttpResponse('Invalid or missing Report ID', status=404)

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Filtered Report'
        
        # Define style for header row
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

        # Add header row to Excel using display names and apply style
        if filtered_data:
            headers = [field_names_mapping.get(field_name, field_name) for field_name in filtered_data[0].keys()]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.style = header_style

        # Add data rows to Excel using values from filtered_data
        for row in filtered_data:
            row_values = [row.get(field_name, '') for field_name in filtered_data[0].keys()]
            sheet.append(row_values)

        # Autofit column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save the workbook to a BytesIO stream
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Prepare the response with Excel file as attachment
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=filtered_report_{report_id}.xlsx'

        return response

class CustomFieldViewset(viewsets.ModelViewSet):
    queryset = Emp_CustomField.objects.all()
    serializer_class = CustomFieldSerializer
    permission_classes = [EmpCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in emp_master._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
    
    # def get_available_fields(self):
    #     emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) ]                
    #     return emp_master_fields

    # @action(detail=False, methods=['get'])
    # def employee_fields(self, request, *args, **kwargs):
    #     available_fields = self.get_available_fields()
    #     return Response({'available_fields': available_fields})


class Emp_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Emp_CustomFieldValue.objects.all()
    serializer_class = Emp_CustomFieldValueSerializer
    permission_classes = [EmpCustomFieldValuePermission]

    
      
class EmpFam_CustomFieldViewset(viewsets.ModelViewSet):
    queryset = EmpFamily_CustomField.objects.all()
    serializer_class = EmpFam_CustomFieldSerializer
    # permission_classes = [EmpFamilyCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in emp_family._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
class Fam_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Fam_CustomFieldValue.objects.all()
    serializer_class = Fam_CustomFieldValueSerializer


class EmpJobHistory_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpJobHistory_CustomField.objects.all()
    serializer_class = EmpJobHistory_Udf_Serializer
    # permission_classes = [EmpJobHistoryCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in EmpJobHistory._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})

class JobHistory_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = JobHistory_CustomFieldValue.objects.all()
    serializer_class = JobHistory_CustomFieldValueSerializer

class EmpQf_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpQualification_CustomField.objects.all()
    serializer_class = Emp_qf_udf_Serializer
    # permission_classes = [EmpQualificationCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in EmpQualification._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})

class Qf_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Qualification_CustomFieldValue.objects.all()
    serializer_class = Qualification_CustomFieldValueSerializer



class EmpDoc_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpDocuments_CustomField.objects.all()
    serializer_class = EmpDocuments_Udf_Serializer
    # permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in Emp_Documents._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
class Doc_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Doc_CustomFieldValue.objects.all()
    serializer_class = DOC_CustomFieldValueSerializer
    # permission_classes = [EmpCustomFieldValuePermission]



class EmpbulkuploadViewSet(viewsets.ModelViewSet):
    queryset = emp_master.objects.all()
    serializer_class = EmpBulkUploadSerializer
    parser_classes = (MultiPartParser, FormParser)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Load workbook and initialize error storage
                    workbook = load_workbook(excel_file)
                    all_errors = {
                        "sheet1_errors": [],
                        "sheet2_errors": []
                    }

                    # Validate presence of sheets
                    sheet1 = workbook.get_sheet_by_name('EmployeeMaster')
                    sheet2 = workbook.get_sheet_by_name('UDF')  # Optional

                    if sheet1 is None or sheet1.max_row == 1:
                        return Response({"error": "Sheet1 is either missing or empty."}, status=400)

                    # Prepare datasets for Sheet1
                    dataset_sheet1 = Dataset()
                    dataset_sheet1.headers = [cell.value for cell in sheet1[1]]
                    for row in sheet1.iter_rows(min_row=2):
                        dataset_sheet1.append([cell.value for cell in row])

                    # Prepare dataset for Sheet2 if it exists
                    dataset_sheet2 = None
                    if sheet2 and sheet2.max_row > 1:
                        dataset_sheet2 = Dataset()
                        dataset_sheet2.headers = [cell.value for cell in sheet2[1]]
                        for row in sheet2.iter_rows(min_row=2):
                            dataset_sheet2.append([str(cell.value) for cell in row])

                    # Resources for import
                    employee_resource = EmployeeResource()
                    custom_field_value_resource = EmpCustomFieldValueResource()

                    # Validate sheet1
                    for row_idx, row in enumerate(dataset_sheet1.dict, start=2):
                        try:
                            employee_resource.before_import_row(row, row_idx=row_idx)
                        except ValidationError as e:
                            all_errors["sheet1_errors"].append({"row": row_idx, "error": str(e)})

                    # Validate sheet2 (if present)
                    if dataset_sheet2:
                        for row_idx, row in enumerate(dataset_sheet2.dict, start=2):
                            try:
                                custom_field_value_resource.before_import_row(row, row_idx=row_idx)
                            except ValidationError as e:
                                all_errors["sheet2_errors"].append({"row": row_idx, "error": str(e)})

                    # Check for errors in both sheets
                    if all_errors["sheet1_errors"] or all_errors["sheet2_errors"]:
                        return Response({"errors": all_errors}, status=400)

                    # Upload data for sheet1
                    with transaction.atomic():
                        employee_result = employee_resource.import_data(dataset_sheet1, dry_run=False, raise_errors=True)

                    # Upload data for sheet2 (if present)
                    if dataset_sheet2:
                        with transaction.atomic():
                            custom_field_value_result = custom_field_value_resource.import_data(dataset_sheet2, dry_run=False, raise_errors=True)
                        return Response({
                            "message": f"{employee_result.total_rows} records created for Sheet1, "
                                    f"{custom_field_value_result.total_rows} records created for Sheet2 successfully"
                        })

                    # If sheet2 is not present
                    return Response({
                        "message": f"{employee_result.total_rows} records created for Sheet1 successfully. Sheet2 was not provided."
                    })

                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)


    
    @action(detail=False, methods=['get'])  # New endpoint for downloading default file
    def download_default_excel_file(self, request):
        # demo_file_path = os.path.join(settings.BASE_DIR,'defaults', 'emp mstr.xlsx')
        demo_file_path = os.path.join(os.path.dirname(__file__), 'defaults', 'Bulkupload_Demo_Sheet.xlsx')
        try:
            with open(demo_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Bulkupload_Demo_Sheet.xlsx"'
                return response
        except FileNotFoundError:
            return Response({"error": "Default demo file not found."}, status=400)

#EMP_FAMILY
class EmpFamViewSet(viewsets.ModelViewSet):
    queryset = emp_family.objects.all()  # Retrieve all instances of emp_family model
    serializer_class = EmpFamSerializer  # Use EmpFamSerializer for serialization
    # permission_classes = [IsAuthenticated]  # Require authentication for access

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:  # Check if user is authenticated
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return emp_family.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # If not a superuser or staff, filter based on emp_id
                return emp_family.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return emp_family.objects.filter(created_by=user)
        return emp_family.objects.none()  # Return an empty queryset if user is not authenticated or does not meet any condition

#EMP_JOB HISTORY
class EmpJobHistoryvSet(viewsets.ModelViewSet):
    queryset = EmpJobHistory.objects.all()
    serializer_class = EmpJobHistorySerializer
    # permission_classes = [IsAuthenticated]
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return EmpJobHistory.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return EmpJobHistory.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return EmpJobHistory.objects.filter(created_by=user)
        return EmpJobHistory.objects.none()
    def get_serializer_context(self):
        return {'request': self.request}
    
#EMP_QUALIFICATION HISTORY
class Emp_QualificationViewSet(viewsets.ModelViewSet):
    queryset = EmpQualification.objects.all()
    serializer_class = Emp_qf_Serializer
    permission_classes = [EmpQualificationCustomFieldPermission]
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return EmpQualification.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return EmpQualification.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return EmpQualification.objects.filter(created_by=user)
        return EmpQualification.objects.none()
    def get_serializer_context(self):
        return {'request': self.request}
    

#EMP_DOCUMENT 
class Emp_DocumentViewSet(viewsets.ModelViewSet):
    queryset = Emp_Documents.objects.all()
    serializer_class = DocumentSerializer
    # permission_classes = [IsAuthenticated]
    def get_serializer_context(self):
        return {'request': self.request}
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return Emp_Documents.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return Emp_Documents.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return Emp_Documents.objects.filter(created_by=user)
        return Emp_Documents.objects.none()
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Check if there are any existing documents with the same emp_id and document_type
        existing_documents = Emp_Documents.objects.filter(
            emp_id=serializer.validated_data['emp_id'],
            document_type=serializer.validated_data['document_type']
        )

        if existing_documents.exists():
            # Deactivate existing documents
            for doc in existing_documents:
                doc.is_active = False
                doc.save()

        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

class Doc_ReportViewset(viewsets.ModelViewSet):
    queryset = Doc_Report.objects.all()
    serializer_class = DocumentReportSerializer
    permission_classes = [DocReportPermission]
    def __init__(self, *args, **kwargs):
        super(Doc_ReportViewset, self).__init__(*args, **kwargs)
        self.ensure_standard_report_exists()

    def get_available_fields(self):
        # Define your available fields logic specific to documents
        excluded_fields = {'id', 'created_at', 'created_by', 'updated_at', 'updated_by', 'emp_sl_no', 'emp_doc_document'}
        included_emp_master_fields = {'emp_first_name', 'emp_active_date', 'emp_branch_id', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id'}
        
        display_names = {
            "emp_id": "Employee ID",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "emp_doc_type": "Document Type",
            "emp_doc_number": "Document Number",
            "emp_doc_issued_date": "Issued Date",
            "emp_doc_expiry_date": "Expiry Date",
            "is_active": "Active",
        }
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        emp_document_fields = [field.name for field in Emp_Documents._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + emp_document_fields}
        return available_fields
    def ensure_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if Doc_Report.objects.filter(file_name='doc_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()   
    def generate_standard_report(self):
        try:
            file_name = 'doc_std_report'
            fields_to_include = self.get_available_fields().keys()
            documents = Emp_Documents.objects.all()

            report_data = self.doc_report_data(fields_to_include, documents)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            Doc_Report.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = Doc_Report.objects.get(file_name='doc_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except Doc_Report.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'])
    def select_document_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return JsonResponse({'available_fields': available_fields})
        
    @action(detail=False, methods=['post'])
    def generate_document_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.data.get('file_name', 'report')
                fields_to_include = request.data.getlist('fields', [])
                # from_date = request.data.get('from_date')
                # to_date = request.data.get('to_date')
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())

            documents = Emp_Documents.objects.all()
            # documents = self.filter_documents_by_date_range(documents, from_date, to_date)

            report_data = self.doc_report_data(fields_to_include, documents)

            if not report_data:
                return JsonResponse({'status': 'error', 'message': 'No data to write into report'})

            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            try:
                with open(file_path, 'w') as file:
                    json.dump(report_data, file, default=str)
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Failed to write file: {str(e)}'})

            try:
                Doc_Report.objects.create(file_name=file_name, report_data=file_name + '.json')
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Failed to save report: {str(e)}'})

            return JsonResponse({'status': 'success', 'file_path': file_path, 'selected_fields_data': fields_to_include})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

    def doc_report_data(self, fields_to_include, documents):
        column_headings = {
            "emp_id": "Employee ID",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "emp_doc_type": "Document Type",
            "emp_doc_number": "Document Number",
            "emp_doc_issued_date": "Issued Date",
            "emp_doc_expiry_date": "Expiry Date",
            "is_active": "Active",
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        emp_document_fields = [field.name for field in Emp_Documents._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        report_data = []
        for document in documents:
            document_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.emp_id, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in emp_document_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                document_data[field] = value
            report_data.append(document_data)

        # print(f"Final report data: {report_data}")
        
        return report_data

    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')
        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        }) 
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = Doc_Report.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except Doc_Report.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'emp_doc_expiry_date' in row and row['emp_doc_expiry_date'] and
            start_date <= datetime.fromisoformat(row['emp_doc_expiry_date']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
      
    @action(detail=False, methods=['post'])
    def generate_doc_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Save selected fields to session
        request.session['selected_fields'] = selected_fields

        # Fetch date-filtered report data from session
        date_filtered_data = request.session.get('date_filtered_data', [])
        print("previosly date filtered ",date_filtered_data)
        
        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            try:
                report = Doc_Report.objects.get(id=report_id)
                report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)
                with open(report_file_path, 'r') as file:
                    report_content = json.load(file)
            except Doc_Report.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Report not found'})

            date_filtered_data = report_content

        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if date_filtered_data:
                selected_fields = list(date_filtered_data[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report
        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,  # Pass filtered data to the frontend
            'unique_values': processed_unique_values,
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        # Extract data from the provided content
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
       
    @action(detail=False, methods=['post'])
    def filter_document_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)

        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]

        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

    # def match_filter_criteria(self, row_data, filter_criteria):
    #     for field, values in filter_criteria.items():
    #         row_value = row_data.get(field, '').strip() if row_data.get(field) else ''
    #         print(f"Checking field {field} with values {values} against row value {row_value}")  # Debugging statement
    #         if row_value not in values:
    #             return False
    #     return True 
    
    @action(detail=False, methods=['get'])
    def generate_excel_report(self, request, *args, **kwargs):
        report_id = request.GET.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        filtered_data = request.session.get('filtered_data')
        if not filtered_data:
            return HttpResponse('No filtered data available', status=400)

        # Mapping of internal field names to display names 
        field_names_mapping = {
            "emp_id": "Employee Code",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "emp_doc_type": "Document Type",
            "emp_doc_number": "Document Number",
            "emp_doc_issued_date": "Issued Date",
            "emp_doc_expiry_date": "Expiry Date",
            "is_active": "Active",
        }

        try:
            report_instance = Doc_Report.objects.get(id=int(report_id))
        except (Doc_Report.DoesNotExist, ValueError):
            return HttpResponse('Invalid or missing Report ID', status=404)

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Filtered Report'
        
        # Define style for header row
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

        # Add header row to Excel using display names and apply style
        if filtered_data:
            headers = [field_names_mapping.get(field_name, field_name) for field_name in filtered_data[0].keys()]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.style = header_style

        # Add data rows to Excel using values from filtered_data
        for row in filtered_data:
            row_values = [row.get(field_name, '') for field_name in filtered_data[0].keys()]
            sheet.append(row_values)

        # Autofit column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save the workbook to a BytesIO stream
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Prepare the response with Excel file as attachment
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=filtered_report_{report_id}.xlsx'
        return response
     
class Bulkupload_DocumentViewSet(viewsets.ModelViewSet):
    queryset = Emp_Documents.objects.all()
    serializer_class = DocBulkuploadSerializer
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Load workbook and initialize error storage
                    workbook = load_workbook(excel_file)
                    all_errors = {
                        "sheet1_errors": [],
                        "sheet2_errors": []
                    }

                    # Validate presence of sheets
                    # sheet1 = workbook.get_sheet_by_name('EmployeeMaster')
                    # sheet2 = workbook.get_sheet_by_name('UDF')  # Optional
                    sheet1 = workbook['DocumentMaster']
                    sheet2 = workbook['Doc-UDF']
                    # sheet2 = workbook.get('UDF', None)  # Optional sheet

                    if sheet1 is None or sheet1.max_row == 1:
                        return Response({"error": "Sheet1 is either missing or empty."}, status=400)

                    # Prepare datasets for Sheet1
                    dataset_sheet1 = Dataset()
                    dataset_sheet1.headers = [cell.value for cell in sheet1[1]]
                    for row in sheet1.iter_rows(min_row=2):
                        dataset_sheet1.append([cell.value for cell in row])

                    # Prepare dataset for Sheet2 if it exists
                    dataset_sheet2 = None
                    if sheet2 and sheet2.max_row > 1:
                        dataset_sheet2 = Dataset()
                        dataset_sheet2.headers = [cell.value for cell in sheet2[1]]
                        for row in sheet2.iter_rows(min_row=2):
                            dataset_sheet2.append([str(cell.value) for cell in row])

                    # Resources for import
                    employee_resource = DocumentResource()
                    custom_field_value_resource = EmpDocumentCustomFieldValueResource()

                    # Validate sheet1
                    for row_idx, row in enumerate(dataset_sheet1.dict, start=2):
                        try:
                            employee_resource.before_import_row(row, row_idx=row_idx)
                        except ValidationError as e:
                            all_errors["sheet1_errors"].append({"row": row_idx, "error": str(e)})

                    # Validate sheet2 (if present)
                    if dataset_sheet2:
                        for row_idx, row in enumerate(dataset_sheet2.dict, start=2):
                            try:
                                custom_field_value_resource.before_import_row(row, row_idx=row_idx)
                            except ValidationError as e:
                                all_errors["sheet2_errors"].append({"row": row_idx, "error": str(e)})

                    # Check for errors in both sheets
                    if all_errors["sheet1_errors"] or all_errors["sheet2_errors"]:
                        return Response({"errors": all_errors}, status=400)
                    

                    # Upload data for sheet1
                    with transaction.atomic():
                        employee_result = employee_resource.import_data(dataset_sheet1, dry_run=False, raise_errors=True)

                    # Upload data for sheet2 (if present)
                    if dataset_sheet2:
                        with transaction.atomic():
                            custom_field_value_result = custom_field_value_resource.import_data(dataset_sheet2, dry_run=False, raise_errors=True)
                        return Response({
                            "message": f"{employee_result.total_rows} records created for Sheet1, "
                                    f"{custom_field_value_result.total_rows} records created for Sheet2 successfully"
                        })

                    # If sheet2 is not present
                    return Response({
                        "message": f"{employee_result.total_rows} records created for Sheet1 successfully. Sheet2 was not provided."
                    })

                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)

class EmplistViewSet(viewsets.ReadOnlyModelViewSet):  # Read-only API (GET only)
    queryset = emp_master.objects.all()
    serializer_class = EmplistSerializer
    permission_classes = [EmployeeListPermission]
    
# EmpLeaveRequest
class EmpLeaveRequestViewSet(viewsets.ModelViewSet):
    queryset = EmpLeaveRequest.objects.all()
    serializer_class = EmpLeaveRequestSerializer
    # permission_classes = [IsAuthenticated]
    def get_serializer_context(self):
        return {'request': self.request}
    

class NotificationViewset(viewsets.ModelViewSet):
    queryset = notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [NotificationPermission]

class EmpMarketSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMarketingSkill.objects.all()
    serializer_class = EmpMarketSkillSerializer 
    permission_classes = [EmployeeMarketingSkillPermission]
 
class EmpPrgrmSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeProgramSkill.objects.all()
    serializer_class = EmpPrgrmSkillSerializer
    permission_classes = [EmployeeProgramSkillPermission]

    
class EmpLangSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeLangSkill.objects.all()
    serializer_class = EmpLangSkillSerializer  
    permission_classes = [EmployeeLangSkillPermission]


class RequestTypeViewset(viewsets.ModelViewSet):
    queryset = RequestType.objects.all()
    serializer_class = RequestTypeSerializer
    permission_classes = [RequestTypePermission]

class EmailTemplateViewset(viewsets.ModelViewSet):
    queryset = EmailTemplate.objects.all()
    serializer_class = EmailTemplateSerializer
    # Custom action to get the placeholders dynamically
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'request': [
                '{{ doc_number }}',
                '{{ request_type }}',
                '{{ reason }}',
                # Add other request-related placeholders here
            ],
            'employee': [
                '{{ doc_number }}',
                '{{ request_type }}',
                '{{ reason }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}'
            ]
        }
        return Response(placeholders)

    # Custom action to fetch the available From and To addresses
    @action(detail=False, methods=['get'], url_path='from-to-addresses')
    def from_to_list(self, request):
        # Fetch active email configurations for "From" addresses
        from_addresses = EmailConfiguration.objects.filter(is_active=True).values_list('email_host_user', flat=True)

        # Fetch employee emails for "To" addresses
        to_addresses = emp_master.objects.all().values_list('emp_personal_email', 'emp_company_email')

        to_list = []
        for emp_personal, emp_company in to_addresses:
            if emp_personal:
                to_list.append(emp_personal)
            if emp_company:
                to_list.append(emp_company)

        return Response({
            'from_addresses': from_addresses,
            'to_addresses': to_list
        })

from django.db import transaction
class GeneralRequestViewset(viewsets.ModelViewSet):
    queryset = GeneralRequest.objects.all()
    serializer_class = GeneralRequestSerializer
    permission_classes =[IsSuperUserOrHasGeneralRequestPermission]
    def perform_create(self, serializer):
        with transaction.atomic():
            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')  # Get manually entered document number

            branch_id = employee.emp_branch_id.id  

            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='general_request',
                    leave_type__isnull=True
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(f"No document numbering configuration found for branch {branch_id} and general request.")

            current_date = timezone.now().date()

            # Validate if the manually entered document number is within the date range
            if document_number:
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError("Document number cannot be assigned outside the valid date range.")
            else:
                # If no document number is entered, generate one automatically
                document_number = doc_config.get_next_number()

            serializer.save(document_number=document_number)

    @action(detail=False, methods=['get'])
    def employee_request_history(self, request):
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'Employee ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        requests = GeneralRequest.get_employee_requests(employee_id)

        # Manually serialize the fields you want
        history_data = []
        for request in requests:
            history_data.append({
                'doc_number': request.doc_number,
                'reason': request.reason,
                'branch': request.branch.branch_name if request.branch else None,
                'request_type': request.request_type.name if request.request_type else None,
                'status': request.status,
                'created_at_date': request.created_at_date,
            })

        return Response(history_data, status=status.HTTP_200_OK)
    

    @action(detail=False, methods=['get'])
    def employee_request_history(self, request):
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'Employee ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        requests = GeneralRequest.get_employee_requests(employee_id)

        # Manually serialize the fields you want
        history_data = []
        for request in requests:
            history_data.append({
                'doc_number': request.doc_number,
                'reason': request.reason,
                'branch': request.branch.branch_name if request.branch else None,
                'request_type': request.request_type.name if request.request_type else None,
                'status': request.status,
                'created_at_date': request.created_at_date,
            })

        return Response(history_data, status=status.HTTP_200_OK)

class ApprovalLevelViewset(viewsets.ModelViewSet):
    queryset = ApprovalLevel.objects.all()
    serializer_class = ApprovalLevelSerializer
    permission_classes = [ApprovalLevelPermission]

class CommonWorkflowViewSet(viewsets.ModelViewSet):
        queryset = CommonWorkflow.objects.all()
        serializer_class  = CommonWorkflowSerializer   

class ApprovalViewset(viewsets.ModelViewSet):
    queryset = Approval.objects.all()
    serializer_class = ApprovalSerializer
    lookup_field = 'pk'
    def get_queryset(self):
        """
        Filter approvals based on the authenticated user.
        """
        user = self.request.user  # Get the logged-in user
        if user.is_superuser:
            return Approval.objects.all()
        return Approval.objects.filter(approver=user)  # Filter approvals assigned to the user

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')  # Get the note from the request
        approval.approve(note=note)
        return Response({'status': 'approved', 'note': note}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')  # Get the note from the request
        approval.reject(note=note)
        return Response({'status': 'rejected', 'note': note}, status=status.HTTP_200_OK)

    
    
class UserNotificationsViewSet(viewsets.ModelViewSet):
    queryset = RequestNotification.objects.all()
    serializer_class = ReqNotifySerializer
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        return RequestNotification.objects.filter(
            Q(recipient_user=user) | Q(recipient_employee__user=user)
        ).order_by('-created_at')  # Fetch only relevant notifications, sorted by latest
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse


class EmailConfigurationViewSet(viewsets.ModelViewSet):
    queryset = EmailConfiguration.objects.all()
    serializer_class = EmailConfigurationSerializer
    lookup_field = 'pk'
    
    def perform_create(self, serializer):
        instance = serializer.save()
        if instance.is_active:
            # Deactivate other configurations
            EmailConfiguration.objects.filter(is_active=True).exclude(pk=instance.pk).update(is_active=False)


    
class GeneralReportViewset(viewsets.ModelViewSet):
    queryset = GeneralRequestReport.objects.all()
    serializer_class = GeneralReportSerializer
    permission_classes = [GeneralRequestReportPermission]
    
    def __init__(self, *args, **kwargs):
        super(GeneralReportViewset, self).__init__(*args, **kwargs)
        self.general_standard_report_exists()

    def get_available_fields(self):
        excluded_fields = {'id', 'created_by'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "branch":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "doc_number": "Document Number",
            "reason": "Reason",
            "total":"Total",
            "request_type": "Request Type",
            "approved": "Approved Request",
            "created_at_date":"Request Date",
           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        general_request_fields = [field.name for field in GeneralRequest._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + general_request_fields}
        return available_fields
    
    @action(detail=False, methods=['get'])
    def select_generalreport_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
       
    @action(detail=False, methods=['post'])
    def generate_general_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            generalreport = GeneralRequest.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,generalreport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            GeneralRequestReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def general_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if GeneralRequestReport.objects.filter(file_name='generalrequest_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'generalrequest_std_report'
            fields_to_include = self.get_available_fields().keys()
            generalreport = GeneralRequest.objects.all()

            report_data = self.generate_report_data(fields_to_include, generalreport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            GeneralRequestReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )
        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = GeneralRequestReport.objects.get(file_name='generalrequest_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except GeneralRequestReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include,generalreport):
        column_headings = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "branch": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "doc_number": "Document Number",
            "reason": "Reason",
            "total":"Total",
            "request_type": "Request Type",
            "approved": "Approved Request",
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        general_request_fields = [field.name for field in GeneralRequest._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = []
        for document in generalreport:
            general_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in general_request_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                general_data[field] = value
            report_data.append(general_data)
        return report_data            

    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        }) 
      
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = GeneralRequestReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except GeneralRequestReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'created_at_date' in row and row['created_at_date'] and
            start_date <= datetime.fromisoformat(row['created_at_date']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
    
    @action(detail=False, methods=['post'])
    def generate_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Fetch previously filtered date data from the `apply_date_filter` method
        date_filtered_data = getattr(self, 'date_filtered_data', [])
        print("previous",date_filtered_data)

        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            report_instance = get_object_or_404(GeneralRequestReport, id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
            date_filtered_data = report_data

        # Default to all fields if no specific fields selected
        if not selected_fields and date_filtered_data:
            selected_fields = list(date_filtered_data[0].keys())

        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {
            field: {'values': values}
            for field, values in unique_values.items()
        }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,
            'unique_values': processed_unique_values,
            # 'column_headings': column_headings
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
    @action(detail=False, methods=['post'])
    def general_filter_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)
        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]
        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

class UpdateESSUserView(APIView):
    def post(self, request, *args, **kwargs):
        selected_employee_ids = request.data.get('selected_employee_ids', [])

        if not selected_employee_ids:
            return Response({'detail': 'At least one selected employee ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

        selected_employees = emp_master.objects.filter(id__in=selected_employee_ids)

        if not selected_employees.exists():
            return Response({'detail': 'No valid employees found.'}, status=status.HTTP_404_NOT_FOUND)

        # Update the SelectedEmpNotify record with the selected employees
        preference, created = SelectedEmpNotify.objects.get_or_create(id=1)
        preference.selected_employees.set(selected_employees)
        preference.save()

        return Response({'detail': 'Selected employees updated successfully.'}, status=status.HTTP_200_OK)


class ESSUserListView(APIView):
    def get(self, request, *args, **kwargs):
        ess_users = emp_master.objects.filter(is_ess=True)
        serializer = EmpSerializer(ess_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class NotificationSettingsViewSet(viewsets.ModelViewSet):
    queryset = NotificationSettings.objects.all()
    serializer_class = NotificationSettingsSerializer

class DocExpEmailTemplateViewset(viewsets.ModelViewSet):
    queryset = DocExpEmailTemplate.objects.all()
    serializer_class = DocExpEmailTemplateSerializer
    
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'employee': [
                '{{ emp_first_name }}',
                '{{ branch }}',
                '{{ reason }}',
                '{{ department }}',
                '{{ designation }}',
                '{{ document_type }}',
                '{{ expiry_date }}',
            ]
        }
        return Response(placeholders)

class EmployeeBankDetailViewset(viewsets.ModelViewSet):
    queryset = EmployeeBankDetail.objects.all()
    serializer_class = EmpBankDetailsSerializer


class EmpBankBulkuploadViewSet(viewsets.ModelViewSet):
    queryset = EmployeeBankDetail.objects.all()
    serializer_class = EmpBankBulkuploadSerializer
    parser_classes = (MultiPartParser, FormParser)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    dataset = Dataset()
                    dataset.load(excel_file.read(), format='xlsx')
                    resource = EmpBankDetailsResource()
                    all_errors = []
                    valid_rows = []
                    with transaction.atomic():
                        for row_idx, row in enumerate(dataset.dict, start=2):
                            row_errors = []
                            try:
                                resource.before_import_row(row, row_idx=row_idx)
                            except ValidationError as e:
                                row_errors.extend([f"Row {row_idx}: {error}" for error in e.messages])
                            if row_errors:
                                all_errors.extend(row_errors)
                            else:
                                valid_rows.append(row)

                    if all_errors:
                        return Response({"errors": all_errors}, status=400)

                    with transaction.atomic():
                        result = resource.import_data(dataset, dry_run=False, raise_errors=True)

                    return Response({"message": f"{result.total_rows} records created successfully"})
                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)
